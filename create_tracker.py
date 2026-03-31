import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Salon Tracker"

# === COLOR PALETTE ===
DARK_BG = "1E293B"       # Slate 800 - header bg
MID_BG = "334155"         # Slate 700
LIGHT_BG = "F8FAFC"       # Slate 50 - alt row
WHITE = "FFFFFF"
TEXT_DARK = "0F172A"      # Slate 900
TEXT_LIGHT = "FFFFFF"
TEXT_MID = "64748B"       # Slate 500

# Brand colors
CLR_DG = "8B5CF6"        # Purple
CLR_NY = "3B82F6"        # Blue  
CLR_OM = "10B981"        # Green
CLR_BIO = "F59E0B"       # Amber
CLR_BKT = "EF4444"       # Red

# Status colors
CLR_PRESENTED = "DBEAFE"  # Light blue
CLR_PURCHASED = "D1FAE5"  # Light green
CLR_NOT = "FEE2E2"        # Light red
CLR_TBC = "FEF3C7"        # Light yellow
CLR_YES = "D1FAE5"        # Light green
CLR_DIGITAL = "E0E7FF"    # Light indigo

# Section header colors
SEC_DG = "7C3AED"         # Purple
SEC_NY = "2563EB"         # Blue
SEC_OM = "059669"         # Green
SEC_FULL = "0F172A"       # Dark
SEC_DGOM = "6D28D9"       # Deep purple
SEC_NYOM = "1D4ED8"       # Deep blue
SEC_BKT = "DC2626"        # Red

# Borders
thin_border = Border(
    left=Side(style='thin', color='E2E8F0'),
    right=Side(style='thin', color='E2E8F0'),
    top=Side(style='thin', color='E2E8F0'),
    bottom=Side(style='thin', color='E2E8F0')
)

# === FONTS ===
title_font = Font(name='Calibri', size=18, bold=True, color=TEXT_DARK)
subtitle_font = Font(name='Calibri', size=10, color=TEXT_MID)
header_font = Font(name='Calibri', size=10, bold=True, color=TEXT_LIGHT)
section_font = Font(name='Calibri', size=11, bold=True, color=TEXT_LIGHT)
salon_font = Font(name='Calibri', size=10, color=TEXT_DARK)
brand_font = Font(name='Calibri', size=10, bold=True, color=TEXT_LIGHT)
status_font = Font(name='Calibri', size=9, bold=True)
small_font = Font(name='Calibri', size=9, color=TEXT_MID)

center = Alignment(horizontal='center', vertical='center', wrap_text=True)
left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

# === COLUMN STRUCTURE ===
# A: Salon Name
# B: Staff Size
# C-G: Brands (DG, NY, O&M, BIO, BKT)
# H: Education
# I: O&M Rewards
# J-K: NPD/Deal 1 (Presented | Purchased)
# L-M: NPD/Deal 2 (Presented | Purchased)
# N-O: NPD/Deal 3 (Presented | Purchased)
# P-Q: NPD/Deal 4 (Presented | Purchased) - spare
# R: Notes

col_widths = {
    'A': 35, 'B': 10, 'C': 8, 'D': 8, 'E': 8, 'F': 8, 'G': 8,
    'H': 12, 'I': 12,
    'J': 14, 'K': 14, 'L': 14, 'M': 14, 'N': 14, 'O': 14, 'P': 14, 'Q': 14,
    'R': 25
}
for col, w in col_widths.items():
    ws.column_dimensions[col].width = w

# === ROW 1: TITLE ===
ws.merge_cells('A1:R1')
ws['A1'] = "SALON CLIENT TRACKER"
ws['A1'].font = title_font
ws['A1'].alignment = Alignment(horizontal='left', vertical='center')
ws.row_dimensions[1].height = 40
for c in range(1, 19):
    ws.cell(row=1, column=c).fill = PatternFill(start_color=WHITE, fill_type='solid')

# === ROW 2: SUBTITLE / LEGEND ===
ws.merge_cells('A2:B2')
ws['A2'] = "Daily tracking · Fill the gaps"
ws['A2'].font = subtitle_font
ws['A2'].alignment = left_align

# Brand legend in row 2
legends = [
    ('C2', 'DG', CLR_DG), ('D2', 'NY', CLR_NY), ('E2', 'O&M', CLR_OM),
    ('F2', 'BIO', CLR_BIO), ('G2', 'BKT', CLR_BKT)
]
for cell_ref, text, color in legends:
    c = ws[cell_ref]
    c.value = text
    c.font = Font(name='Calibri', size=9, bold=True, color=TEXT_LIGHT)
    c.fill = PatternFill(start_color=color, fill_type='solid')
    c.alignment = center

# Status legend
status_legends = [
    ('H2', '✓ = YES', CLR_YES),
    ('I2', 'D = DIGITAL', CLR_DIGITAL),
]
for cell_ref, text, color in status_legends:
    c = ws[cell_ref]
    c.value = text
    c.font = Font(name='Calibri', size=8, bold=True, color=TEXT_DARK)
    c.fill = PatternFill(start_color=color, fill_type='solid')
    c.alignment = center

status_legends2 = [
    ('J2', 'P = PRESENTED', CLR_PRESENTED),
    ('K2', '$ = PURCHASED', CLR_PURCHASED),
    ('L2', 'TBC', CLR_TBC),
    ('M2', '✗ = NOT YET', CLR_NOT),
]
for cell_ref, text, color in status_legends2:
    c = ws[cell_ref]
    c.value = text
    c.font = Font(name='Calibri', size=8, bold=True, color=TEXT_DARK)
    c.fill = PatternFill(start_color=color, fill_type='solid')
    c.alignment = center

ws.row_dimensions[2].height = 22

# === ROW 3: BLANK SPACER ===
ws.row_dimensions[3].height = 6
for c in range(1, 19):
    ws.cell(row=3, column=c).fill = PatternFill(start_color=WHITE, fill_type='solid')

# === ROW 4: NPD/DEAL HEADERS (the changeable ones) ===
row = 4
ws.row_dimensions[row].height = 28

# Fixed headers
fixed_headers = [
    ('A', 'SALON'), ('B', 'STAFF'), ('C', 'DG'), ('D', 'NY'), ('E', 'O&M'),
    ('F', 'BIO'), ('G', 'BKT'), ('H', 'EDU'), ('I', 'REWARDS')
]
for col_letter, text in fixed_headers:
    c = ws[f'{col_letter}{row}']
    c.value = text
    c.font = header_font
    c.fill = PatternFill(start_color=DARK_BG, fill_type='solid')
    c.alignment = center
    c.border = thin_border

# Brand-specific header colors
brand_header_colors = {'C': CLR_DG, 'D': CLR_NY, 'E': CLR_OM, 'F': CLR_BIO, 'G': CLR_BKT}
for col_letter, color in brand_header_colors.items():
    ws[f'{col_letter}{row}'].fill = PatternFill(start_color=color, fill_type='solid')

# NPD/Deal paired columns (these are the ones that change bi-monthly)
deal_headers = [
    ('J', 'K', 'DG BERET DEAL 20'),
    ('L', 'M', 'BIOTECH 18'),
    ('N', 'O', 'HBC DEALS'),
    ('P', 'Q', 'NY DEAL'),
]
for p_col, s_col, deal_name in deal_headers:
    # Merge the pair for the deal name
    p_idx = openpyxl.utils.column_index_from_string(p_col)
    s_idx = openpyxl.utils.column_index_from_string(s_col)
    ws.merge_cells(start_row=row, start_column=p_idx, end_row=row, end_column=s_idx)
    c = ws.cell(row=row, column=p_idx)
    c.value = deal_name
    c.font = Font(name='Calibri', size=9, bold=True, color=TEXT_LIGHT)
    c.fill = PatternFill(start_color=MID_BG, fill_type='solid')
    c.alignment = center

# Notes header
ws[f'R{row}'] = 'NOTES'
ws[f'R{row}'].font = header_font
ws[f'R{row}'].fill = PatternFill(start_color=DARK_BG, fill_type='solid')
ws[f'R{row}'].alignment = center

# === ROW 5: Sub-headers for P/S under each deal ===
row = 5
ws.row_dimensions[row].height = 20

# Repeat fixed headers (smaller)
for col_letter, text in fixed_headers:
    c = ws[f'{col_letter}{row}']
    c.value = ''
    c.fill = PatternFill(start_color=DARK_BG, fill_type='solid')
    c.border = thin_border
for col_letter, color in brand_header_colors.items():
    ws[f'{col_letter}{row}'].fill = PatternFill(start_color=color, fill_type='solid')

# P / S sub-headers
for p_col, s_col, _ in deal_headers:
    for col, label in [(p_col, 'PRESENT'), (s_col, 'PURCH')]:
        c = ws[f'{col}{row}']
        c.value = label
        c.font = Font(name='Calibri', size=8, bold=True, color=TEXT_LIGHT)
        c.fill = PatternFill(start_color=MID_BG, fill_type='solid')
        c.alignment = center
        c.border = thin_border

ws[f'R{row}'] = ''
ws[f'R{row}'].fill = PatternFill(start_color=DARK_BG, fill_type='solid')

# === SALON DATA ===
salons = [
    # (section_name, section_color, salons_list)
    # Each salon: (name, staff, DG, NY, OM, BIO, BKT, edu, rewards, deals_presented, deals_purchased, notes)
    # deals: list of (deal_idx, presented_status, purchased_status) — deal_idx 0-3
    
    ("DUN GUD ONLY", SEC_DG, [
        ("Pipsqueek In Saigon Stepney", "4+", True, False, False, True, False, "YES", "", [], ""),
        ("Jade Hairs (CLOSED MONDAY)", "SOLO", True, False, False, True, False, "DIGITAL", "", [], ""),
        ("Just Peachy Hairdressing", "1-2", True, False, False, False, False, "YES", "", [], ""),
        ("Thirteen (closed Mon / Tue)", "1-2", True, False, False, False, False, "", "", [], ""),
        ("Jacky Boy", "2-3", True, False, False, True, False, "YES", "", [], ""),
        ("Fluffee (Open 10-4pm)", "SOLO", True, False, False, False, False, "DIGITAL", "", [], ""),
        ("Mane & Marsden", "SOLO", True, False, False, False, False, "", "", [], ""),
    ]),
    ("NINE YARDS ONLY", SEC_NY, [
        ("Kabuki Hair", "4+", False, True, False, False, True, "YES", "", [], ""),
    ]),
    ("O&M ONLY", SEC_OM, [
        ("ALLURE", "10+", False, False, True, False, True, "", "", [], ""),
        ("C00590 Oria Hair Studio", "4+", False, False, True, True, True, "YES", "", [], ""),
        ("LILA ROSE", "2-3", False, False, True, False, False, "YES", "", [], ""),
        ("C13587 Where She Goes", "2-3", False, False, True, True, False, "YES", "TBC", [], ""),
        ("C06679 GLOW HAIR AND MAKEUP BY ROCHELLE", "1-2", False, False, True, False, False, "DIGITAL", "YES", [], ""),
        ("C15990 Denika Santillo Hair Artistry", "SOLO", False, False, True, False, False, "YES", "", [], ""),
        ("C00203 Del & May Hair Studio", "1-2", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C12022 Asri Hair Salon", "1-2", False, False, True, True, False, "YES", "YES", [], ""),
        ("C06289 Kristy Harrison", "2-3", False, False, True, True, False, "YES", "", [], ""),
        ("C05416 Udara", "2-3", False, False, True, False, False, "YES", "", [], ""),
        ("C08429 The Meraki Lounge", "1-2", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C16798 Serenity Hair and Beauty Lounge", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C18755 Hairhouse Marion", "", False, False, True, False, False, "YES", "", [], ""),
        ("C11286 Hair By Fiona Marshall", "SOLO", False, False, True, True, True, "YES", "", [], ""),
        ("C20633 Hair House West Lakes", "", False, False, True, False, False, "YES", "", [], ""),
        ("C23070 Full Circle Hair Society", "4+", False, False, True, False, False, "YES", "", [], ""),
        ("C16523 Bernadette Kelly", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C11583 Talisha Lynch", "SOLO", False, False, True, False, False, "YES", "", [], ""),
        ("C15674 The Little Hair Parlour", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C17851 Stone Salon", "SOLO", False, False, True, False, False, "", "", [], ""),
        ("C11743 GEORGIA RAÈ HAIR", "1-2", False, False, True, False, False, "YES", "", [], ""),
        ("C20436 Style Hut", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C01944 Rebecca Rosa Hair (Ensemble Salon)", "SOLO", False, False, True, False, False, "YES", "", [], ""),
        ("C00176 Comb Culture", "4+", False, False, True, False, False, "DIGITAL", "YES", [], ""),
        ("C15956 S & M Hair", "2-3", False, False, True, False, False, "DIGITAL", "TBC", [], ""),
        ("C00925 Hair Poetry", "4+", False, False, True, False, True, "DIGITAL", "YES", [], ""),
        ("C13286 SCALPT & CO", "4+", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C12163 BeyondBlonde by Denise Mayberg", "1-2", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C00780 Studio 8", "1-2", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C08708 Intricate Hair", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C12596 Browz & Cutz", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C15113 Salt Hair Artistry", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
        ("C00577 Nyuba Hair", "SOLO", False, False, True, False, False, "DIGITAL", "", [], ""),
    ]),
    ("FULL CONCEPT HBC", SEC_FULL, [
        ("C00722 Scizzor Lounge", "4+", True, True, True, True, False, "YES", "YES", [], ""),
        ("C06854 Aveta Hair and Body", "2-3", True, True, True, True, True, "YES", "TBC", [], ""),
        ("C00194 Dare Hair", "4+", True, True, True, True, True, "YES", "YES", [], ""),
        ("C12158 SALTY BLONDES", "SOLO", True, True, True, True, False, "YES", "", [], ""),
    ]),
    ("DUN GUD / O&M", SEC_DGOM, [
        ("C12792 Daisy & Co The Salon", "4+", True, False, True, True, False, "YES", "TBC", [], ""),
        ("C01850 Hair Crush @ The Beach", "SOLO", True, False, True, False, False, "", "", [], ""),
        ("C03850 The Collective Hair Bar", "2-3", True, False, True, True, True, "YES", "TBC", [], ""),
        ("C09803 Hairstablishment", "1-2", True, False, True, True, False, "YES", "", [], ""),
        ("C01166 Muse The Hair Room", "2-3", True, False, True, True, True, "YES", "", [], ""),
        ("C11589 Hair by Ellie Clancy", "SOLO", True, False, True, False, False, "YES", "", [], ""),
        ("C18410 Emily Allegretto", "SOLO", True, False, True, False, False, "", "", [], ""),
        ("C08430 Benny Mae. The Salon", "SOLO", True, False, True, True, True, "YES", "", [], ""),
        ("C23275 BYDEMIROSE", "SOLO", True, False, True, False, False, "YES", "", [], ""),
    ]),
    ("NINE YARDS / O&M", SEC_NYOM, [
        ("C00958 Salon Super Store", "4+", False, True, True, False, False, "", "", [], ""),
        ("C03697 Emily Lauren Hair", "SOLO", False, True, True, True, True, "YES", "", [], ""),
    ]),
    ("BKT ONLY", SEC_BKT, [
        ("Khrome Hair", "", False, False, False, False, True, "", "", [], ""),
        ("Glamazon Hair NT", "", False, False, False, False, True, "", "", [], ""),
        ("Lavish Hair Studio", "", False, False, False, False, True, "", "", [], ""),
        ("Hair on Grange", "", False, False, False, False, True, "", "", [], ""),
    ]),
]

# PS S+C Target data (from right side of original)
ps_targets = [
    ("C00590 Oria Hair Studio", "5TH MARCH"),
    ("LILA ROSE", "5TH MARCH"),
    ("C15990 Denika Santillo Hair Artistry", "TBC"),
    ("C06289 Kristy Harrison", "19TH MARCH - TRAINING"),
    ("C05416 Udara", "5TH MARCH"),
    ("C16798 Serenity Hair and Beauty Lounge", "MARIA TRAINING 10TH MARCH"),
    ("C11286 Hair By Fiona Marshall", "1ST APRIL MEETING"),
    ("C20633 Hair House West Lakes", "5TH MARCH"),
    ("C23070 Full Circle Hair Society", "TBC"),
    ("C11743 GEORGIA RAÈ HAIR", "3RD MARCH TRAINING"),
    ("C01944 Rebecca Rosa Hair (Ensemble Salon)", "17TH MARCH TRAINING"),
    ("C12158 SALTY BLONDES", "18TH MARCH MEETING"),
    ("C03850 The Collective Hair Bar", "12TH MARCH MEETING"),
    ("C09803 Hairstablishment", "24TH MARCH MEETING"),
    ("C01166 Muse The Hair Room", "11TH MARCH"),
    ("C00958 Salon Super Store", "11TH MARCH"),
]

# Build a lookup for PS target notes
ps_lookup = {name: date for name, date in ps_targets}

current_row = 6

def write_section_header(ws, row, section_name, color):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=18)
    c = ws.cell(row=row, column=1)
    c.value = f"  {section_name}"
    c.font = section_font
    c.fill = PatternFill(start_color=color, fill_type='solid')
    c.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[row].height = 26
    # Fill all cells in the row
    for col in range(2, 19):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color=color, fill_type='solid')

def write_salon_row(ws, row, salon_data, is_alt):
    name, staff, dg, ny, om, bio, bkt, edu, rewards, deals, notes = salon_data
    bg_color = LIGHT_BG if is_alt else WHITE
    bg_fill = PatternFill(start_color=bg_color, fill_type='solid')
    
    # A: Name
    c = ws.cell(row=row, column=1, value=name)
    c.font = salon_font
    c.fill = bg_fill
    c.alignment = left_align
    c.border = thin_border
    
    # B: Staff
    c = ws.cell(row=row, column=2, value=staff)
    c.font = Font(name='Calibri', size=9, color=TEXT_MID)
    c.fill = bg_fill
    c.alignment = center
    c.border = thin_border
    
    # C-G: Brands
    brand_cols = [(3, dg, CLR_DG), (4, ny, CLR_NY), (5, om, CLR_OM), (6, bio, CLR_BIO), (7, bkt, CLR_BKT)]
    for col, has_brand, brand_color in brand_cols:
        c = ws.cell(row=row, column=col)
        if has_brand:
            c.value = "●"
            c.font = Font(name='Calibri', size=14, bold=True, color=brand_color)
            c.fill = bg_fill
        else:
            c.value = ""
            c.fill = bg_fill
        c.alignment = center
        c.border = thin_border
    
    # H: Education
    c = ws.cell(row=row, column=8)
    if edu == "YES":
        c.value = "✓"
        c.font = Font(name='Calibri', size=11, bold=True, color="059669")
        c.fill = PatternFill(start_color=CLR_YES, fill_type='solid')
    elif edu == "DIGITAL":
        c.value = "D"
        c.font = Font(name='Calibri', size=10, bold=True, color="4F46E5")
        c.fill = PatternFill(start_color=CLR_DIGITAL, fill_type='solid')
    else:
        c.value = ""
        c.fill = bg_fill
    c.alignment = center
    c.border = thin_border
    
    # I: Rewards
    c = ws.cell(row=row, column=9)
    if rewards == "YES":
        c.value = "✓"
        c.font = Font(name='Calibri', size=11, bold=True, color="059669")
        c.fill = PatternFill(start_color=CLR_YES, fill_type='solid')
    elif rewards == "TBC":
        c.value = "TBC"
        c.font = Font(name='Calibri', size=9, bold=True, color="92400E")
        c.fill = PatternFill(start_color=CLR_TBC, fill_type='solid')
    else:
        c.value = ""
        c.fill = bg_fill
    c.alignment = center
    c.border = thin_border
    
    # J-Q: Deal columns (4 deals × 2 cols each) - leave blank for manual entry
    for col in range(10, 18):
        c = ws.cell(row=row, column=col)
        c.value = ""
        c.fill = bg_fill
        c.alignment = center
        c.border = thin_border
    
    # R: Notes (include PS target date if exists)
    note_text = notes
    if name in ps_lookup:
        note_text = f"PS: {ps_lookup[name]}" + (f" | {notes}" if notes else "")
    c = ws.cell(row=row, column=18)
    c.value = note_text
    c.font = small_font
    c.fill = bg_fill
    c.alignment = left_align
    c.border = thin_border
    
    ws.row_dimensions[row].height = 24

# Write all sections
for section_name, section_color, salon_list in salons:
    write_section_header(ws, current_row, section_name, section_color)
    current_row += 1
    for i, salon in enumerate(salon_list):
        write_salon_row(ws, current_row, salon, i % 2 == 1)
        current_row += 1
    # Small gap between sections
    ws.row_dimensions[current_row].height = 4
    for col in range(1, 19):
        ws.cell(row=current_row, column=col).fill = PatternFill(start_color=WHITE, fill_type='solid')
    current_row += 1

# === SUMMARY ROW at bottom ===
current_row += 1
ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=18)
c = ws.cell(row=current_row, column=1)
total_salons = sum(len(s[2]) for s in salons)
c.value = f"  TOTAL SALONS: {total_salons}  |  Updated: March 2026  |  💡 TIP: Change deal names in row 4 when new NPD launches"
c.font = Font(name='Calibri', size=10, italic=True, color=TEXT_MID)
c.fill = PatternFill(start_color=LIGHT_BG, fill_type='solid')
c.alignment = Alignment(horizontal='left', vertical='center')
for col in range(2, 19):
    ws.cell(row=current_row, column=col).fill = PatternFill(start_color=LIGHT_BG, fill_type='solid')

# === FREEZE PANES ===
ws.sheet_view.zoomScale = 100
ws.freeze_panes = 'A6'  # Freeze headers

# === PRINT SETUP ===
ws.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

# Save
output_path = '/Users/steve/agent/projects/salon-tracker/Salon_Client_Tracker.xlsx'
wb.save(output_path)
print(f"Created: {output_path}")
print(f"Total salons: {total_salons}")
